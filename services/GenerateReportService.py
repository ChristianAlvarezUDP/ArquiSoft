import sys
import socket
import threading
import sqlite3
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def dict_factory(cursor, row):
    fields = [column[0] for column in cursor.description]
    return {key: value for key, value in zip(fields, row)}


def request(bus_ip, bus_port, service_name, message):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as client_socket:
        client_socket.connect((bus_ip, bus_port))
        request = f"{service_name}:{message}"
        client_socket.sendall(request.encode('utf-8'))
        response = client_socket.recv(8192)

        return response.decode('utf-8')
    

def filtrar_auditorias(auditoria, form_id):
    date_format = "%Y-%m-%d %H:%M:%S"
    fecha = datetime.strptime(auditoria['marca_temporal'], date_format)

    now = datetime.now()
    time_diff = abs(now - fecha)

    print(form_id, auditoria['id_formulario'])

    return auditoria['id_formulario'] == form_id and time_diff <= timedelta(hours=24)


def generar_reporte(form_id):
    filename = "reporte.xlsx"

    field_aliases = {
        'tipo': 'Tipo Auditoria',
        'marca_temporal': 'Fecha'
    }

    blocked_fields = ['id', 'fecha', 'formulario', 'id_formulario']

    query = {
        'comando': 'get_all_auditorias'
    }

    response = request('127.0.0.1', 5000, 'AuditoriaService.py', json.dumps(query))
    auditorias_data = list(filter(lambda auditoria: filtrar_auditorias(auditoria, form_id), json.loads(response)['auditorias']))

    print(auditorias_data)

    if not auditorias_data:
        return False

    query = {
        'comando': 'get_form',
        'body': {
            'id_grupo_campos': form_id
        }
    }

    response = request('127.0.0.1', 5000, 'GestionFormulariosService.py', json.dumps(query))
    form_data = json.loads(response)

    headers = list(auditorias_data[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = form_data['body']['formulario']['nombre']

    ws["A1"] = "Auditoria: " + form_data['body']['formulario']['nombre']
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    header_fill = PatternFill(start_color="81d41a", end_color="81d41a", fill_type="solid")
    border_style = Side(border_style="thin", color="000000")

    # N
    cell = ws.cell(row=2, column=2, value="Nº")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

    offset = 0
    for col_idx, header in enumerate(headers, start=3):
        if header in blocked_fields:
            offset -= 1
            continue

        value = field_aliases[header] if header in field_aliases else header.capitalize().replace("_", " ")
        cell = ws.cell(row=2, column=col_idx + offset, value=value)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

    for col_idx, pregunta in enumerate(form_data['body']['preguntas'], start=len(headers) + 3):
        cell = ws.cell(row=2, column=col_idx + offset, value=pregunta['titulo'])
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

    for row_idx in range(len(auditorias_data)):
        cell = ws.cell(row=row_idx + 3, column=2, value=row_idx + 1)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

    for row_idx, row_data in enumerate(auditorias_data, start=3):
        offset = 0

        for col_idx, header in enumerate(headers, start=3):
            if header in blocked_fields:
                offset -= 1
                continue

            cell = ws.cell(row=row_idx, column=col_idx + offset, value=row_data[header])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

        query = {
            'comando': 'get_auditoria',
            'auditoria_id': row_data['id']
        }

        response = request('127.0.0.1', 5000, 'AuditoriaService.py', json.dumps(query))
        respuestas = json.loads(response)['auditoria']['respuestas']

        for col_idx, pregunta in enumerate(form_data['body']['preguntas'], start=len(headers) + 3):
            respuesta = list(filter(lambda respuesta: respuesta['titulo'] == pregunta['titulo'], respuestas))

            respuesta = '' if not respuesta else respuesta[0]['valor']

            cell = ws.cell(row=row_idx, column=col_idx + offset, value=respuesta)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)


    ws.column_dimensions[get_column_letter(1)].width = 20

    for col_idx, header in enumerate(headers, start=2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(filename)

    return True


def get_auditorias_by_auditor(idAuditor):
    query = '''
        SELECT a.id, a.marca_temporal, a.fecha, gc.nombre AS grupo_campos_nombre, 
               b.n_interno AS bus_interno, ta.nombre AS tipo_auditoria_nombre, 
               au.nombre AS auditor_nombre
        FROM auditoria AS a
        JOIN grupo_campos AS gc ON a.id_grupo_campos = gc.id
        JOIN bus AS b ON a.id_bus = b.id
        JOIN tipo_auditoria AS ta ON a.id_tipo_auditoria = ta.id
        JOIN auditor AS au ON a.id_auditor = au.id
        WHERE a.id_auditor = ?
    '''
    try:
        conn = sqlite3.connect('sqlite/arqui.db')
        conn.row_factory = dict_factory
        cursor = conn.cursor()
        cursor.execute(query, (idAuditor,))
        result = cursor.fetchall()
    
    except sqlite3.Error as e:
        print(f"Error fetching data: {e}")
        return json.dumps({"error": str(e)})
    finally:
        cursor.close()
        conn.close()

    return json.dumps(result)

def service_worker(service_name, host, port):
    print(f"{service_name} starting on {host}:{port}")
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as server_socket:
        server_socket.bind((host, port))
        server_socket.listen()
        while True:
            client_socket, _ = server_socket.accept()
            data = client_socket.recv(1024).decode('utf-8')
            print(f"{service_name} recibio: {data}")
            
            data = json.loads(data)
            response = ""

            if data["comando"] == "generar_reporte":
                result = generar_reporte(data["id_form"])

                if result:
                    response = {
                        "status": "correct"
                    }
                else:
                    response = {
                        "status": "error",
                        "error": "No hay auditorias de este formulario dentro de 24 horas."
                    }

                response = json.dumps(response)
            
            if data["comando"] == "AuditoriasPorAuditor":
                response = get_auditorias_by_auditor(data["body"]["id_auditor"])
                if response == []:
                    response = "No se encontraron auditorias"        

            print(f"{service_name} responde: {response}")
            client_socket.sendall(response.encode('utf-8'))
            client_socket.close()

if __name__ == '__main__':
    threading.Thread(target=service_worker, args=(sys.argv[0].split('/')[-1], '127.0.0.1', int(sys.argv[1]))).start()
